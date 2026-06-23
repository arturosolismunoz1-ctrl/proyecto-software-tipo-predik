"""
Prueba funcional — Ecatepec, Estado de México.

Flujo completo a través de la API del sistema:
  1. Login  -> JWT
  2. ETL    -> POST /api/v1/admin/etl/denue/run  (descarga DENUE real de Ecatepec)
  3. Zona   -> POST /api/v1/zona/concentracion-comercial  (consulta el cubo)
  4. Raw DB -> consulta directa a raw_data.denue_establishments para detalle
  5. Excel  -> genera reporte con resumen + detalle de papelerías

Uso:
  python backend/scripts/prueba_funcional_ecatepec.py
  python backend/scripts/prueba_funcional_ecatepec.py --output resultados/ecatepec.xlsx
"""
import argparse
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración ──────────────────────────────────────────────────────────────
API_BASE = os.getenv("API_BASE", "http://localhost:8000/api/v1")
EMAIL    = os.getenv("PREDIK_EMAIL", "admin@predik.local")
PASSWORD = os.getenv("PREDIK_PASSWORD", "dev_password_admin")

# Polígono que cubre Ecatepec de Morelos, Estado de México
ECATEPEC_POLYGON = {
    "type": "Polygon",
    "coordinates": [[
        [-99.10, 19.55],
        [-98.94, 19.55],
        [-98.94, 19.76],
        [-99.10, 19.76],
        [-99.10, 19.55],
    ]],
}

# SCIAN: 451110 = Comercio al por menor en tiendas de papelería y artículos de escritorio
# También buscamos por nombre para capturar variantes
PAPELERIA_KEYWORDS = ["papel", "papelería", "papeleria", "librería", "libreria", "útiles", "utiles"]
PAPELERIA_SCIAN_PREFIXES = ["4511", "451"]


# ── Helpers de estilo Excel ───────────────────────────────────────────────────

def _header_fill(r, g, b):
    return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_header_row(ws, headers, row=1, fill=None):
    fill = fill or _header_fill(30, 80, 140)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

def _write_data_row(ws, row, values, alt=False):
    alt_fill = _header_fill(240, 245, 255) if alt else _header_fill(255, 255, 255)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="center")
        cell.fill = alt_fill
        cell.border = _thin_border()

def _autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


# ── Paso 1: Login ─────────────────────────────────────────────────────────────

def login() -> str:
    print("1/5 -> Login en la API...")
    r = requests.post(f"{API_BASE}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=10)
    r.raise_for_status()
    token = r.json()["access_token"]
    print(f"     [OK] JWT obtenido")
    return token


# ── Paso 2: ETL DENUE para Ecatepec ──────────────────────────────────────────

def run_etl(token: str) -> dict:
    print("2/5 -> Ejecutando ETL DENUE — papelerias Ecatepec (estado=15, keyword=papeleria)...")
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.post(
        f"{API_BASE}/admin/etl/inegi_denue/run",
        json={
            "estado": "15",          # Estado de México
            "keyword": "papeleria",  # Filtra solo papelerías (INEGI DENUE v1 no soporta municipio)
            "max_records": 2500,     # Max de la API — ~295 caen en Ecatepec tras filtro bbox
            "h3_resolution": 9,
            "polygon": ECATEPEC_POLYGON,
        },
        headers=headers,
        timeout=180,
    )
    if r.status_code != 200:
        print(f"     [!] ETL retornó {r.status_code}: {r.text[:300]}")
        return {}
    result = r.json()
    extracted = result.get("extracted", 0)
    loaded    = result.get("loaded", 0)
    aggregated = result.get("aggregated", 0)
    print(f"     [OK] {extracted} extraídos | {loaded} en raw_data | {aggregated} celdas H3")
    return result


# ── Paso 3: Concentración comercial ──────────────────────────────────────────

def get_concentracion(token: str) -> dict:
    print("3/5 -> Consultando concentración comercial de Ecatepec...")
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.post(
        f"{API_BASE}/zona/concentracion-comercial",
        json={"geometry": ECATEPEC_POLYGON},
        headers=headers,
        timeout=30,
    )
    if r.status_code == 404:
        print("     [!] Zona sin cobertura en el cubo (ETL puede no haber cargado datos)")
        return {}
    r.raise_for_status()
    result = r.json()
    print(f"     [OK] {result.get('total_establecimientos', 0)} establecimientos totales en la zona")
    print(f"     [OK] {len(result.get('por_categoria', []))} categorías SCIAN detectadas")
    print(f"     [OK] analysis_id: {result.get('analysis_id')}")
    return result


# ── Paso 4: Consulta directa a la DB para detalle de papelerías ──────────────

def get_papelerias_detalle() -> list[dict]:
    print("4/5 -> Consultando raw_data para detalle de papelerías...")
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parents[2] / ".env")

    from sqlalchemy import select, or_, func
    from app.db import SessionLocal
    from app.models.raw_data import DenueEstablishment

    db = SessionLocal()
    try:
        # Filtrar por SCIAN o por nombre
        scian_filters = [
            DenueEstablishment.codigo_scian.like(f"{p}%")
            for p in PAPELERIA_SCIAN_PREFIXES
        ]
        name_filters = [
            func.lower(DenueEstablishment.nombre).like(f"%{kw}%")
            for kw in PAPELERIA_KEYWORDS
        ]
        clase_filters = [
            func.lower(DenueEstablishment.clase_actividad).like(f"%{kw}%")
            for kw in PAPELERIA_KEYWORDS
        ]

        rows = db.execute(
            select(DenueEstablishment).where(
                or_(*scian_filters, *name_filters, *clase_filters)
            )
        ).scalars().all()

        result = []
        for r in rows:
            lat = lon = None
            if r.geom is not None:
                try:
                    from sqlalchemy import func as sqlfunc
                    coords = db.execute(
                        select(
                            func.ST_X(func.ST_GeomFromWKB(DenueEstablishment.geom)),
                            func.ST_Y(func.ST_GeomFromWKB(DenueEstablishment.geom)),
                        ).where(DenueEstablishment.id == r.id)
                    ).one_or_none()
                    if coords:
                        lon, lat = coords
                except Exception:
                    pass

            result.append({
                "clee": r.clee or "",
                "nombre": r.nombre or "",
                "clase_actividad": r.clase_actividad or "",
                "codigo_scian": r.codigo_scian or "",
                "estrato_personal": r.estrato_personal or "",
                "entidad": r.entidad or "",
                "municipio": r.municipio or "",
                "colonia": r.colonia or "",
                "cp": r.cp or "",
                "lat": lat,
                "lon": lon,
            })

        print(f"     [OK] {len(result)} papelerías encontradas en raw_data")
        return result

    finally:
        db.close()


# ── Paso 5: Generar Excel ─────────────────────────────────────────────────────

def generar_excel(
    concentracion: dict,
    papelerias: list[dict],
    output_path: str,
) -> None:
    print(f"5/5 -> Generando Excel: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen Ecatepec"
    ws_res.row_dimensions[1].height = 40

    zona = concentracion.get("zona", {})
    por_cat = concentracion.get("por_categoria", [])
    total = concentracion.get("total_establecimientos", 0)
    ancla = concentracion.get("negocios_ancla", [])

    # Título
    ws_res.merge_cells("A1:F1")
    title_cell = ws_res["A1"]
    title_cell.value = f"Concentración Comercial — Ecatepec de Morelos, Edo. Méx."
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _header_fill(30, 80, 140)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata
    meta = [
        ("Entidad", zona.get("entidad", "N/D")),
        ("Municipio", zona.get("municipio", "N/D")),
        ("Área analizada (km²)", zona.get("area_km2", 0)),
        ("Total establecimientos", total),
        ("Negocios ancla detectados", len(ancla)),
        ("Papelerías encontradas", len(papelerias)),
        ("Fecha del análisis", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Fuente de datos", "INEGI DENUE (API real)"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws_res.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_res.cell(row=i, column=2, value=value)

    # Tabla de categorías
    ws_res.cell(row=12, column=1, value="Top categorías SCIAN en la zona").font = Font(bold=True, size=11)
    _write_header_row(ws_res, ["#", "Categoría SCIAN", "Código SCIAN", "Establecimientos", "% del total"], row=13)

    for i, cat in enumerate(por_cat[:20], start=14):
        pct = round(cat["cantidad"] / total * 100, 1) if total > 0 else 0
        _write_data_row(ws_res, i, [
            i - 13,
            cat["categoria"],
            cat["codigo_scian"],
            cat["cantidad"],
            f"{pct}%",
        ], alt=(i % 2 == 0))

    _autofit(ws_res)

    # ── Hoja 2: Papelerías ─────────────────────────────────────────────────────
    ws_pap = wb.create_sheet("Papelerías")
    ws_pap.row_dimensions[1].height = 35

    ws_pap.merge_cells("A1:J1")
    t = ws_pap["A1"]
    t.value = f"Papelerías en Ecatepec de Morelos — {len(papelerias)} establecimientos"
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = _header_fill(20, 120, 80)
    t.alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "CLEE", "Nombre del establecimiento", "Clase de actividad",
        "SCIAN", "Personal ocupado", "Entidad", "Municipio",
        "Colonia", "C.P.", "Latitud", "Longitud",
    ]
    _write_header_row(ws_pap, headers, row=2, fill=_header_fill(20, 120, 80))

    for i, p in enumerate(papelerias, start=3):
        _write_data_row(ws_pap, i, [
            p["clee"],
            p["nombre"],
            p["clase_actividad"],
            p["codigo_scian"],
            p["estrato_personal"],
            p["entidad"],
            p["municipio"],
            p["colonia"],
            p["cp"],
            p["lat"],
            p["lon"],
        ], alt=(i % 2 == 0))

    _autofit(ws_pap)

    # ── Hoja 3: Todas las categorías ──────────────────────────────────────────
    ws_cat = wb.create_sheet("Todas las categorías")
    ws_cat.merge_cells("A1:E1")
    t2 = ws_cat["A1"]
    t2.value = "Desglose completo por categoría SCIAN — Ecatepec"
    t2.font = Font(bold=True, size=12, color="FFFFFF")
    t2.fill = _header_fill(100, 60, 140)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    _write_header_row(ws_cat, ["#", "Categoría", "Código SCIAN", "Establecimientos", "% del total"], row=2,
                      fill=_header_fill(100, 60, 140))

    for i, cat in enumerate(por_cat, start=3):
        pct = round(cat["cantidad"] / total * 100, 2) if total > 0 else 0
        _write_data_row(ws_cat, i, [
            i - 2,
            cat["categoria"],
            cat["codigo_scian"],
            cat["cantidad"],
            f"{pct}%",
        ], alt=(i % 2 == 0))

    _autofit(ws_cat)

    wb.save(output_path)
    print(f"     [OK] Excel guardado en: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Prueba funcional Ecatepec — Predik Geo")
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parents[2] / "resultados" / "ecatepec_papelerias.xlsx"),
    )
    parser.add_argument("--skip-etl", action="store_true", help="Omitir ETL (usar datos ya en DB)")
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  PRUEBA FUNCIONAL — Ecatepec, Estado de México")
    print("  Sistema: Predik Geo | API:", API_BASE)
    print("=" * 60)

    token = login()

    if not args.skip_etl:
        run_etl(token)
    else:
        print("2/5 -> [ETL omitido por --skip-etl]")

    concentracion = get_concentracion(token)
    papelerias = get_papelerias_detalle()
    generar_excel(concentracion, papelerias, args.output)

    print("\n" + "=" * 60)
    print("  RESULTADO FINAL")
    print("=" * 60)
    print(f"  Total establecimientos en Ecatepec : {concentracion.get('total_establecimientos', 0)}")
    print(f"  Papelerías encontradas             : {len(papelerias)}")
    print(f"  Archivo Excel                      : {args.output}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
