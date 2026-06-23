"""
Servicio de generacion de reportes geoespaciales.

Dado un poligono + capas de busqueda, el servicio:
  1. Corre ETL por cada capa (INEGI DENUE)
  2. Consulta celdas H3 del cubo
  3. Consulta puntos individuales por capa
  4. Clasifica hexagonos (densidad o oportunidad competitiva)
  5. Genera KMZ o Excel en memoria y devuelve bytes

Soporta cualquier combinacion: papelerias, restaurantes, farmacias,
competencia entre cadenas, etc.
"""
import io
import json
import zipfile
from datetime import date
from typing import Any, Dict, List, Literal, Optional
from xml.sax.saxutils import escape

import h3
import openpyxl
from geoalchemy2 import functions as geo_funcs
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.etl.denue import DenueETL
from app.models.raw_data import AgebDemographics, AgebGeometry, DenueEstablishment
from app.services.zona_analysis import calculate_commercial_concentration


# ── Paleta de colores ──────────────────────────────────────────────────────────
# Formato KML: AABBGGRR (alpha, blue, green, red)

_COLOR_PUNTO = {
    "red":    "FF0000FF",
    "green":  "FF00CC00",
    "blue":   "FFFF6600",
    "yellow": "FF00EEFF",
    "orange": "FF0066FF",
    "purple": "FFCC0099",
    "cyan":   "FFFFFF00",
    "pink":   "FF9900FF",
}

_COLOR_ICONO_URL = {
    "red":    "http://maps.google.com/mapfiles/kml/paddle/red-circle.png",
    "green":  "http://maps.google.com/mapfiles/kml/paddle/grn-circle.png",
    "blue":   "http://maps.google.com/mapfiles/kml/paddle/blu-circle.png",
    "yellow": "http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png",
    "orange": "http://maps.google.com/mapfiles/kml/paddle/orange-circle.png",
    "purple": "http://maps.google.com/mapfiles/kml/paddle/purple-circle.png",
    "cyan":   "http://maps.google.com/mapfiles/kml/paddle/ltblu-circle.png",
    "pink":   "http://maps.google.com/mapfiles/kml/paddle/pink-circle.png",
}

# Iconos tipo estrella para marcas específicas
_COLOR_ICONO_STAR_URL = {
    "red":    "http://maps.google.com/mapfiles/kml/paddle/red-stars.png",
    "green":  "http://maps.google.com/mapfiles/kml/paddle/grn-stars.png",
    "blue":   "http://maps.google.com/mapfiles/kml/paddle/blu-stars.png",
    "yellow": "http://maps.google.com/mapfiles/kml/paddle/ylw-stars.png",
    "orange": "http://maps.google.com/mapfiles/kml/paddle/wht-stars.png",
    "purple": "http://maps.google.com/mapfiles/kml/paddle/purple-stars.png",
    "cyan":   "http://maps.google.com/mapfiles/kml/paddle/ltblu-stars.png",
    "pink":   "http://maps.google.com/mapfiles/kml/paddle/pink-stars.png",
}

# Colores hexagonos por clasificacion de oportunidad
_OPT_FILL = {
    "ALTA":       "CC00C800",   # verde oscuro
    "MEDIA_ALTA": "CC44E844",   # verde claro
    "MEDIA":      "CC00CCDD",   # amarillo
    "BAJA":       "CC999999",   # gris
    "SATURADA":   "CC0000CC",   # rojo
}

_OPT_LABEL = {
    "ALTA":       "ALTA OPORTUNIDAD",
    "MEDIA_ALTA": "OPORTUNIDAD MEDIA-ALTA",
    "MEDIA":      "OPORTUNIDAD MEDIA",
    "BAJA":       "BAJA ACTIVIDAD",
    "SATURADA":   "SATURADA",
}


# ── 1. ETL por capa ────────────────────────────────────────────────────────────

async def run_etl_capas(
    db: Session,
    capas: List[Dict[str, Any]],
    polygon: Dict[str, Any],
    max_records: int,
    h3_resolution: int,
) -> List[Dict[str, int]]:
    """Ejecuta ETL DENUE para cada capa. Retorna stats por capa."""
    resultados = []
    for capa in capas:
        etl = DenueETL()
        stats = await etl.run(
            db,
            resolution=h3_resolution,
            estado=capa["estado"],
            keyword=capa["keyword"],
            max_records=max_records,
            polygon=polygon,
        )
        resultados.append({"capa": capa["label"], **stats})
    return resultados


# ── 2. Consulta de puntos por capa ─────────────────────────────────────────────

def query_puntos_capa(
    db: Session,
    polygon: Dict[str, Any],
    keyword: str,
    scian_prefix: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Establecimientos individuales dentro del poligono que coincidan con keyword."""
    polygon_json = json.dumps(polygon)
    kw = keyword.lower().strip()

    stmt = select(
        DenueEstablishment.nombre,
        DenueEstablishment.clase_actividad,
        DenueEstablishment.codigo_scian,
        DenueEstablishment.estrato_personal,
        DenueEstablishment.colonia,
        DenueEstablishment.municipio,
        func.ST_X(DenueEstablishment.geom).label("lon"),
        func.ST_Y(DenueEstablishment.geom).label("lat"),
    ).where(
        DenueEstablishment.geom.isnot(None),
        geo_funcs.ST_Intersects(
            DenueEstablishment.geom,
            geo_funcs.ST_GeomFromGeoJSON(polygon_json),
        ),
    )

    filtros_texto = [
        func.lower(DenueEstablishment.nombre).like(f"%{kw}%"),
        func.lower(DenueEstablishment.clase_actividad).like(f"%{kw}%"),
    ]
    if scian_prefix:
        filtros_texto.append(DenueEstablishment.codigo_scian.like(f"{scian_prefix}%"))
    stmt = stmt.where(or_(*filtros_texto))

    rows = db.execute(stmt).all()
    return [
        {
            "nombre":          r.nombre or "",
            "clase_actividad": r.clase_actividad or "",
            "codigo_scian":    r.codigo_scian or "",
            "estrato_personal": r.estrato_personal or "",
            "colonia":         r.colonia or "",
            "municipio":       r.municipio or "",
            "lat":             float(r.lat),
            "lon":             float(r.lon),
        }
        for r in rows if r.lat and r.lon
    ]


# ── 3. Consulta AGEBs reales ──────────────────────────────────────────────────

def query_agebs_en_poligono(
    db: Session,
    polygon: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Consulta AGEBs que intersectan el polígono con conteo de establecimientos DENUE.
    Hace LEFT JOIN con demographics para datos de población.
    Retorna lista de dicts con geom (GeoJSON), cvegeo, conteos.
    Retorna [] si la tabla ageb_geometries está vacía (aún no se ha cargado el MGN).
    """
    polygon_json = json.dumps(polygon)

    stmt = (
        select(
            AgebGeometry.cvegeo,
            AgebGeometry.nom_mun,
            AgebGeometry.nom_loc,
            AgebGeometry.ambito,
            func.ST_AsGeoJSON(AgebGeometry.geom).label("geom"),
            AgebDemographics.pobtot,
            AgebDemographics.p_0a14,
            AgebDemographics.p_15a64,
            AgebDemographics.p_65ymas,
            AgebDemographics.graproes,
            func.count(DenueEstablishment.id).label("num_establecimientos"),
        )
        .outerjoin(AgebDemographics, AgebGeometry.cvegeo == AgebDemographics.cvegeo)
        .outerjoin(
            DenueEstablishment,
            geo_funcs.ST_Within(DenueEstablishment.geom, AgebGeometry.geom),
        )
        .where(
            geo_funcs.ST_Intersects(
                AgebGeometry.geom,
                geo_funcs.ST_GeomFromGeoJSON(polygon_json),
            )
        )
        .group_by(
            AgebGeometry.cvegeo,
            AgebGeometry.nom_mun,
            AgebGeometry.nom_loc,
            AgebGeometry.ambito,
            AgebGeometry.geom,
            AgebDemographics.pobtot,
            AgebDemographics.p_0a14,
            AgebDemographics.p_15a64,
            AgebDemographics.p_65ymas,
            AgebDemographics.graproes,
        )
    )

    rows = db.execute(stmt).all()
    return [
        {
            "cvegeo":              r.cvegeo,
            "nom_mun":             r.nom_mun or "",
            "nom_loc":             r.nom_loc or "",
            "ambito":              r.ambito or "Urbana",
            "geom":                r.geom or "",
            "pobtot":              r.pobtot or 0,
            "p_0a14":              r.p_0a14 or 0,
            "p_15a64":             r.p_15a64 or 0,
            "p_65ymas":            r.p_65ymas or 0,
            "graproes":            float(r.graproes) if r.graproes else 0.0,
            "cantidad":            int(r.num_establecimientos),
            # intensidad se calcula después normalizando
            "intensidad":          0.0,
        }
        for r in rows
    ]


def _normalizar_intensidad(agebs: List[Dict]) -> List[Dict]:
    max_c = max((a["cantidad"] for a in agebs), default=1) or 1
    for a in agebs:
        a["intensidad"] = a["cantidad"] / max_c
    return agebs


# ── 3. Clasificacion de hexagonos ──────────────────────────────────────────────

def clasificar_por_densidad(hexagonos: List[Dict]) -> List[Dict]:
    """Colorea hexagonos segun cantidad: verde -> amarillo -> naranja -> rojo."""
    max_c = max((h["cantidad"] for h in hexagonos), default=1)
    result = []
    for h in hexagonos:
        ratio = 0.0 if max_c <= 1 else min(1.0, (h["cantidad"] - 1) / (max_c - 1))
        if ratio < 0.33:
            t = ratio / 0.33
            r, g, b = int(t * 255), 200, 0
        elif ratio < 0.66:
            t = (ratio - 0.33) / 0.33
            r, g, b = 255, int(200 - t * 128), 0
        else:
            t = (ratio - 0.66) / 0.34
            r, g, b = 255, int(72 - t * 72), 0
        color = f"CC{b:02X}{g:02X}{r:02X}"
        label = f"{h['cantidad']} establecimiento{'s' if h['cantidad'] != 1 else ''}"
        result.append({**h, "color": color, "label": label})
    return result


def clasificar_por_oportunidad(
    hexagonos: List[Dict],
    capas_con_puntos: List[Dict],
    h3_resolution: int,
) -> List[Dict]:
    """
    Clasifica hexagonos segun presencia de cadenas y densidad comercial.

    Logica:
      - Celda con TODAS las cadenas presentes -> SATURADA
      - Celda con ALGUNAS cadenas             -> MEDIA (zona activa, con competencia)
      - Celda sin cadenas + alta densidad     -> ALTA OPORTUNIDAD
      - Celda sin cadenas + media densidad    -> MEDIA-ALTA
      - Celda sin cadenas + baja densidad     -> BAJA
    """
    # Para cada capa, calcular en que celdas H3 cae
    capa_cells: Dict[str, set] = {}
    for capa in capas_con_puntos:
        cells = {
            h3.latlng_to_cell(p["lat"], p["lon"], h3_resolution)
            for p in capa.get("puntos", [])
            if p.get("lat") and p.get("lon")
        }
        capa_cells[capa["label"]] = cells

    n_capas = len(capas_con_puntos)

    # Quantiles de intensidad para rangos relativos
    intensidades = sorted(h["intensidad"] for h in hexagonos)
    n = len(intensidades)
    q33 = intensidades[n // 3] if n > 2 else 0
    q66 = intensidades[2 * n // 3] if n > 2 else 0

    result = []
    for h in hexagonos:
        n_presentes = sum(
            1 for cells in capa_cells.values()
            if h["h3_index"] in cells
        )

        if n_capas > 1 and n_presentes == n_capas:
            nivel = "SATURADA"
        elif n_presentes > 0:
            nivel = "MEDIA"
        else:
            if h["intensidad"] >= q66:
                nivel = "ALTA"
            elif h["intensidad"] >= q33:
                nivel = "MEDIA_ALTA"
            else:
                nivel = "BAJA"

        color = _OPT_FILL.get(nivel, "CC888888")
        label = _OPT_LABEL.get(nivel, nivel)
        result.append({**h, "color": color, "label": label, "nivel": nivel})

    return result


# ── 3b. Clasificacion por poder adquisitivo (AGEBs) ───────────────────────────

_PA_FILL = {
    "PREMIUM":    "CC005500",   # verde oscuro intenso
    "MEDIO_ALTO": "CC22AA22",   # verde medio
    "MEDIO":      "CC00AAFF",   # ámbar
    "BAJO":       "CC888888",   # gris
}

_PA_LABEL = {
    "PREMIUM":    "Zona Premium (escolaridad alta)",
    "MEDIO_ALTO": "Zona Medio-Alta",
    "MEDIO":      "Zona Media",
    "BAJO":       "Zona Baja actividad",
}


def clasificar_por_poder_adquisitivo(agebs: List[Dict]) -> List[Dict]:
    """
    Clasifica AGEBs por poder adquisitivo usando grado promedio de escolaridad (graproes).
    Fuente: INEGI Censo 2020 — graproes = años promedio de escolaridad población 15+.
      >= 12 años (preparatoria+) → PREMIUM
      >= 9  años (secundaria)    → MEDIO_ALTO
      >= 6  años (primaria)      → MEDIO
      <  6  años                 → BAJO
    """
    result = []
    for a in agebs:
        gpa = float(a.get("graproes", 0) or 0)
        if gpa >= 12:
            nivel = "PREMIUM"
        elif gpa >= 9:
            nivel = "MEDIO_ALTO"
        elif gpa >= 6:
            nivel = "MEDIO"
        else:
            nivel = "BAJO"
        result.append({
            **a,
            "color": _PA_FILL[nivel],
            "label": _PA_LABEL[nivel],
            "nivel": nivel,
        })
    return result


# ── 4. Generacion KMZ ──────────────────────────────────────────────────────────

def _geojson_to_kml_coords(geom_str: str) -> str:
    try:
        geom = json.loads(geom_str)
    except Exception:
        return ""
    if geom["type"] == "Polygon":
        ring = geom["coordinates"][0]
    elif geom["type"] == "MultiPolygon":
        ring = geom["coordinates"][0][0]
    else:
        return ""
    return " ".join(f"{c[0]:.6f},{c[1]:.6f},0" for c in ring)


def _kml_estilo_punto(sid: str, color_nombre: str, icon_type: str = "circle") -> List[str]:
    icon_color = _COLOR_PUNTO.get(color_nombre, "FF0000FF")
    if icon_type == "star":
        icon_url = _COLOR_ICONO_STAR_URL.get(color_nombre,
                       "http://maps.google.com/mapfiles/kml/paddle/ylw-stars.png")
        scale = "1.1"
    else:
        icon_url = _COLOR_ICONO_URL.get(color_nombre,
                       "http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png")
        scale = "0.9"
    return [
        f'  <Style id="{sid}">',
        f'    <IconStyle><color>{icon_color}</color><scale>{scale}</scale>',
        f'      <Icon><href>{icon_url}</href></Icon>',
        f'    </IconStyle>',
        f'    <LabelStyle><scale>0</scale></LabelStyle>',
        f'  </Style>',
    ]


def _kml_estilo_hex(sid: str, color: str) -> List[str]:
    return [
        f'  <Style id="{sid}">',
        f'    <LineStyle><color>88111111</color><width>0.3</width></LineStyle>',
        f'    <PolyStyle><color>{color}</color></PolyStyle>',
        f'  </Style>',
    ]


def generar_kmz(
    nombre: str,
    hexagonos: List[Dict],
    capas_con_puntos: List[Dict],
) -> bytes:
    """Genera bytes de un KMZ con hexagonos + puntos por capa."""
    colores_usados_hex = {h["color"] for h in hexagonos}

    L: List[str] = []
    L.append('<?xml version="1.0" encoding="UTF-8"?>')
    L.append('<kml xmlns="http://www.opengis.net/kml/2.2">')
    L.append('<Document>')
    L.append(f'  <name>{escape(nombre)}</name>')

    total_puntos = sum(len(c.get("puntos", [])) for c in capas_con_puntos)
    L.append(
        f'  <description><![CDATA['
        f'{total_puntos} establecimientos en {len(capas_con_puntos)} capa(s) | '
        f'{len(hexagonos)} celdas H3<br/>'
        f'Fuente: INEGI DENUE — {date.today()}'
        f']]></description>'
    )

    # Estilos puntos
    for capa in capas_con_puntos:
        sid = f"capa_{capa['label'].lower().replace(' ', '_')[:20]}"
        capa["_style_id"] = sid
        L.extend(_kml_estilo_punto(sid, capa.get("color", "blue"), capa.get("icon", "circle")))

    # Estilos hexagonos (uno por color unico)
    for color in colores_usados_hex:
        L.extend(_kml_estilo_hex(f"hex_{color}", color))

    # Carpeta zonas (AGEBs o H3 según disponibilidad)
    usa_agebs = any("cvegeo" in h for h in hexagonos)
    carpeta_label = "AGEBs INEGI" if usa_agebs else "Hexagonos H3"
    if hexagonos:
        L.append('  <Folder>')
        L.append(f'    <name>{carpeta_label}</name>')
        L.append(f'    <description>{len(hexagonos)} zonas</description>')
        for h in sorted(hexagonos, key=lambda x: x.get("intensidad", 0)):
            coords = _geojson_to_kml_coords(h.get("geom", ""))
            if not coords:
                continue
            h_label = h.get("label", "")
            if "cvegeo" in h:
                desc = (
                    f"AGEB: <b>{h.get('cvegeo','')}</b><br/>"
                    f"Municipio: {escape(h.get('nom_mun',''))}<br/>"
                    f"Establecimientos: {h.get('cantidad', 0)}<br/>"
                    f"Poblacion: {h.get('pobtot', 0):,}<br/>"
                    f"Escolaridad prom: {h.get('graproes', 0):.1f} anos<br/>"
                    f"Clasificacion: <b>{h_label}</b>"
                )
            else:
                desc = (
                    f"Clasificacion: <b>{h_label}</b><br/>"
                    f"Establecimientos: {h.get('cantidad', 0)}<br/>"
                    f"Intensidad: {round(h.get('intensidad', 0) * 100, 2)}%<br/>"
                    f"H3 index: {h.get('h3_index', '')}"
                )
            color_sid = f"hex_{h['color']}"
            L += [
                '    <Placemark>',
                f'      <name>{escape(h_label)}</name>',
                f'      <description><![CDATA[{desc}]]></description>',
                f'      <styleUrl>#{color_sid}</styleUrl>',
                '      <Polygon><tessellate>1</tessellate>',
                '        <outerBoundaryIs><LinearRing>',
                f'          <coordinates>{coords}</coordinates>',
                '        </LinearRing></outerBoundaryIs>',
                '      </Polygon>',
                '    </Placemark>',
            ]
        L.append('  </Folder>')

    # Carpeta por cada capa
    for capa in capas_con_puntos:
        puntos = capa.get("puntos", [])
        L.append('  <Folder>')
        L.append(f'  <name>{escape(capa["label"])} ({len(puntos)})</name>')
        L.append(f'  <description>Keyword: {escape(capa["keyword"])} | Estado: {escape(capa["estado"])}</description>')
        for p in puntos:
            desc = (
                f"Clase: {escape(p['clase_actividad'])}<br/>"
                f"SCIAN: {escape(p['codigo_scian'])}<br/>"
                f"Personal: {escape(p['estrato_personal'])}<br/>"
                f"Colonia: {escape(p['colonia'])}"
            )
            L += [
                '    <Placemark>',
                f'      <name>{escape(p["nombre"])}</name>',
                f'      <description><![CDATA[{desc}]]></description>',
                f'      <styleUrl>#{capa["_style_id"]}</styleUrl>',
                '      <Point>',
                f'        <coordinates>{p["lon"]:.6f},{p["lat"]:.6f},0</coordinates>',
                '      </Point>',
                '    </Placemark>',
            ]
        L.append('  </Folder>')

    L.append('</Document>')
    L.append('</kml>')

    kml_bytes = '\n'.join(L).encode('utf-8')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('doc.kml', kml_bytes)
    return buf.getvalue()


# ── 5. Generacion Excel ────────────────────────────────────────────────────────

def _excel_header(ws, headers, row=1, rgb=(30, 80, 140)):
    fill = PatternFill("solid", fgColor=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")
    side = Side(style="thin", color="CCCCCC")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def _excel_row(ws, row, values, alt=False):
    fill = PatternFill("solid", fgColor="F0F5FF" if alt else "FFFFFF")
    side = Side(style="thin", color="DDDDDD")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(vertical="center")
        c.border = border


def _autofit(ws):
    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 50)


def generar_excel(
    nombre: str,
    hexagonos: List[Dict],
    capas_con_puntos: List[Dict],
) -> bytes:
    """Genera bytes de un Excel con hoja de resumen + una hoja por capa."""
    wb = openpyxl.Workbook()

    # ── Resumen ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = nombre
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1E508C")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    meta = [
        ("Fecha del analisis", str(date.today())),
        ("Capas de busqueda", len(capas_con_puntos)),
        ("Total establecimientos", sum(len(c.get("puntos", [])) for c in capas_con_puntos)),
        ("Celdas H3 analizadas", len(hexagonos)),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.cell(row=8, column=1, value="Detalle por capa").font = Font(bold=True, size=11)
    _excel_header(ws, ["Capa", "Keyword", "Estado", "Establecimientos"], row=9)
    for i, capa in enumerate(capas_con_puntos, 10):
        _excel_row(ws, i, [
            capa["label"], capa["keyword"], capa["estado"],
            len(capa.get("puntos", [])),
        ], alt=(i % 2 == 0))
    _autofit(ws)

    # ── Hoja por capa ─────────────────────────────────────────────────────────
    for capa in capas_con_puntos:
        puntos = capa.get("puntos", [])
        ws2 = wb.create_sheet(capa["label"][:31])  # Excel max 31 chars
        ws2.merge_cells("A1:H1")
        t2 = ws2["A1"]
        t2.value = f"{capa['label']} — {len(puntos)} establecimientos"
        t2.font = Font(bold=True, size=12, color="FFFFFF")
        t2.fill = PatternFill("solid", fgColor="1E508C")
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30

        hdrs = ["Nombre", "Clase de actividad", "SCIAN", "Personal", "Colonia", "Municipio", "Latitud", "Longitud"]
        _excel_header(ws2, hdrs, row=2)
        for i, p in enumerate(puntos, 3):
            _excel_row(ws2, i, [
                p["nombre"], p["clase_actividad"], p["codigo_scian"],
                p["estrato_personal"], p["colonia"], p["municipio"],
                p["lat"], p["lon"],
            ], alt=(i % 2 == 0))
        _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Funcion principal ──────────────────────────────────────────────────────────

async def generar_reporte(
    db: Session,
    organization_id: str,
    nombre: str,
    polygon: Dict[str, Any],
    capas: List[Dict[str, Any]],
    formato: Literal["kmz", "excel"],
    clasificacion_hexagonos: Literal["densidad", "oportunidad", "poder_adquisitivo"],
    max_records: int,
    h3_resolution: int,
    ejecutar_etl: bool,
) -> bytes:
    """
    Punto de entrada principal del servicio de reportes.

    Parametros de cada capa:
      keyword   - termino de busqueda para INEGI DENUE
      label     - nombre visible en el reporte
      color     - red | green | blue | yellow | orange | purple | cyan | pink
      estado    - codigo de entidad INEGI (2 digitos, ej. "15" = EdoMex, "14" = Jalisco)
      scian_prefix - (opcional) prefijo SCIAN para filtro adicional
    """
    # 1. ETL
    if ejecutar_etl:
        await run_etl_capas(db, capas, polygon, max_records, h3_resolution)

    # 2. Zonas: intentar AGEBs reales primero, caer en H3 si no hay MGN cargado
    hexagonos_raw: List[Dict] = []
    usa_agebs = False
    try:
        agebs = query_agebs_en_poligono(db, polygon)
        if agebs:
            hexagonos_raw = _normalizar_intensidad(agebs)
            usa_agebs = True
    except Exception:
        pass

    if not usa_agebs:
        try:
            result = calculate_commercial_concentration(db, organization_id, polygon)
            hexagonos_raw = result.get("celdas_heatmap", [])
        except ValueError:
            pass

    # 3. Puntos por capa
    capas_con_puntos = []
    for capa in capas:
        puntos = query_puntos_capa(
            db,
            polygon,
            capa["keyword"],
            scian_prefix=capa.get("scian_prefix"),
        )
        capas_con_puntos.append({**capa, "puntos": puntos})

    # 4. Clasificar hexagonos / AGEBs
    if not hexagonos_raw:
        hexagonos: List[Dict] = []
    elif clasificacion_hexagonos == "poder_adquisitivo" and usa_agebs:
        hexagonos = clasificar_por_poder_adquisitivo(hexagonos_raw)
    elif clasificacion_hexagonos == "oportunidad" and len(capas_con_puntos) >= 1:
        hexagonos = clasificar_por_oportunidad(hexagonos_raw, capas_con_puntos, h3_resolution)
    else:
        hexagonos = clasificar_por_densidad(hexagonos_raw)

    # 5. Generar salida
    if formato == "kmz":
        return generar_kmz(nombre, hexagonos, capas_con_puntos)
    else:
        return generar_excel(nombre, hexagonos, capas_con_puntos)
